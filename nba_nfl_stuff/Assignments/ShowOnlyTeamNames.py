from nba_nfl_stuff import GlobalFunctions
import openpyxl


def remove_rows(workbook, rows):
    for i in reversed(rows):
        print("Deleteing row " + str(i))
        workbook.delete_rows(i)


def do_assignment(document):
    wrkbk = openpyxl.load_workbook("Output/assignment_1_" + document + ".xlsx")
    sh = wrkbk.active

    team_column_pos = GlobalFunctions.get_column_pos(sh, "team")

    # We will use the win (W) column to confirm that the row is legit.
    # If it's a number it's good, if not it's bad
    win_column_pos = GlobalFunctions.get_column_pos(sh, "W")

    rows_to_be_removed = []
    # iterate through excel and display data
    for i in range(2, sh.max_row + 1):

        team_name = sh.cell(row=i, column=team_column_pos).value
        wins = sh.cell(row=i, column=win_column_pos).value

        if type(wins) != int:
            print(team_name + " <- Will be removed")
            rows_to_be_removed.append(i)
        else:
            print(team_name)

    remove_rows(sh, rows_to_be_removed)

    wrkbk.save(filename="output/Assignment_2_" + document + ".xlsx")
