from ToolBox import ToolScript

import openpyxl
import CleanTeamNames

wrkbk = openpyxl.load_workbook("../Resources/nfl.xlsx")
sh = wrkbk.active

team_column_pos = ToolScript.getTeamColumnPos(sh)

rows_to_be_removed = []


def removeRows(rows):

    for i in reversed(rows):
        print(i)
        sh.delete_rows(i)

# iterate through excel and display data
for i in range(2, sh.max_row + 1):

    teamName = sh.cell(row=i, column=team_column_pos).value

    # Headers, luckily for us, are very similar and easy to identify and remove
    if teamName[0:3] == "NFC" or teamName[0:3] == "AFC":
        print(CleanTeamNames.clean_text(teamName) + "<-- TO BE REMOVED!!")
        rows_to_be_removed.append(i)

    else:
        print(CleanTeamNames.clean_text(teamName) + "<- Stays")

removeRows(rows_to_be_removed)

wrkbk.save(filename="../output/Assignment_2.xlsx")