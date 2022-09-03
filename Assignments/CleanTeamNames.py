import GlobalFunctions

import re
import openpyxl


def clean_text(text):
    # Removing anything that isn't a normal letter or number
    # Maybe we could remove parenthesis and numbers here, but just in case a team name actually contains a number
    # we'll stick to only removing number that exist between parenthesis
    cleantext = re.sub("[^A-Za-z0-9().\s]+", '', text)

    # If there is a start parenthesis, we'll remove it and anything behind it
    if "(" in cleantext:
        cleantext = cleantext[0:cleantext.index("(")]

    return cleantext


def do_assignment(document):
    # load excel with its path
    wrkbk = openpyxl.load_workbook("Resources/" + document + ".xlsx")
    sh = wrkbk.active


    team_column_pos = GlobalFunctions.get_column_pos(sh, "team")

    # iterate through excel and display data
    for i in range(1, sh.max_row + 1):
        print("\n")
        print("Row ", i)

        # Printing original value
        cell_content = sh.cell(row=i, column=team_column_pos).value
        print("Original String -> " + cell_content)

        # Printing clean value
        cleanCellContent = clean_text(cell_content)
        print('Clean String -> ' + cleanCellContent)

        # Replaceing the old value with the corected one
        sh.cell(row=i, column=team_column_pos).value = cleanCellContent

    # Saving file
    wrkbk.save(filename="output/assignment_1_" + document + ".xlsx")
