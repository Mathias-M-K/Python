import openpyxl

from openpyxl.workbook import Workbook

pattern_one_column = 1
pattern_one_row = 1

pattern_two_column = 13
pattern_two_row = 1


def is_pattern_one(current_line, prev_line):
    return current_line == prev_line + 1


def is_pattern_two(current_line, prev_line):
    return current_line == prev_line - 1


def write_to_pattern_one(workbook, data):
    print("Pattern one!")
    global pattern_one_row
    global pattern_one_column

    for dataEntry in data:
        y = 0
        for column_value in dataEntry:
            print("row:", pattern_one_row, "column:", pattern_one_column + y, "value:", column_value)
            workbook.cell(row=pattern_one_row, column=pattern_one_column + y).value = column_value
            y += 1

        pattern_one_row += 1

    pattern_one_row += 1


def write_to_pattern_two(workbook, data):
    print("Pattern Two!")
    global pattern_two_row
    global pattern_two_column

    for dataEntry in data:
        y = 0
        for column_value in dataEntry:
            print("row:", pattern_two_row, "column:", pattern_two_column + y, "value:", column_value)
            workbook.cell(row=pattern_two_row, column=pattern_two_column + y).value = column_value
            y += 1

        pattern_two_row += 1

    pattern_two_row += 1


def save_workbook(workbook):
    workbook.save("output_my.xlsx")


def row_to_tuple(sheet, row, number_of_columns):
    values = []
    for i in range(1, number_of_columns):
        values.append(sheet.cell(row=row, column=i).value)

    return tuple(values)


def do_stuff():
    # opening workbook and getting the active sheet. I guess it's always number one? No idea why
    workbook = openpyxl.load_workbook("Book2.xlsx")
    sheet = workbook.active

    workbook_output = Workbook()
    sheet_out = workbook_output.active

    # Creating a list to store our pattern
    pattern_content = []

    # To recognize a pattern, we need to know the content of previous row and if we are already working a pattern
    prev_row = None
    pattern_active = False
    pattern_type = 0
    for i in range(1, sheet.max_row + 2):

        # if it's the first row, we set prev_row and skip
        if i == 1:
            prev_row = row_to_tuple(sheet, i, 11)
            continue

        index = sheet.cell(row=i, column=1).value

        # If current and prev values match any of our patterns, we save the value in our list
        if is_pattern_one(index, prev_row[0]) or is_pattern_two(index, prev_row[0]):
            pattern_active = True
            pattern_content.append(prev_row)
            pattern_type = 1 if is_pattern_one(index, prev_row[0]) else 2
        else:
            if pattern_active:
                pattern_content.append(prev_row)
                print("Pattern found", pattern_content)

                if pattern_type == 1:
                    write_to_pattern_one(sheet_out, pattern_content)
                else:
                    write_to_pattern_two(sheet_out, pattern_content)

                pattern_content = []
                pattern_active = False
                pattern_type = 0

        prev_row = row_to_tuple(sheet, i, 11)

    save_workbook(workbook_output)


# Does stuff
do_stuff()
